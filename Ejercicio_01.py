import os
import openpyxl 
from docxtpl import DocxTemplate

def valida_opcion(numero):
    assert numero>0 and numero<3, 'Opción inválida'
    assert int(numero), 'La opcion ingresada no es un digito'

def combinar_correspondencia():

    book = openpyxl.load_workbook('DatosClientes.xlsx', data_only=True)
    word = DocxTemplate('BANCOABCtemplate.docx') 

    hoja = book.active

    listado = []
    cuentas = []

    for fila in hoja.iter_rows(min_row=2,min_col=1, max_col=hoja.max_column-3):
        empleado = ([celda.value for celda in fila])
        if empleado not in listado:
            listado.append(empleado)
    
    for fila in hoja.iter_rows(min_row=1,min_col=1, max_col=hoja.max_column-1):
            cuenta = ([celda.value for celda in fila])
            cuentas.append(cuenta)


    for i in range(0,len(listado)):    
        cuentas_asociadas = []
        saldos_cuenta = []
        for j, detalle in enumerate(cuentas):
            if listado[i][0] in detalle:
                cuentas_asociadas.append(cuentas[j][1])
                saldos_cuenta.append(cuentas[j][2])

        context = {
            'nombre': listado[i][0],
            'cuentas': cuentas_asociadas,
            'saldos': saldos_cuenta,
            'total': sum(saldos_cuenta)
        }

        word.render(context)
        word.save(f"nuevodoc_generado{i+1}.docx")
    print(f'Generación de documentación completada!')


def menu():
    opcion=0
    while opcion!=2:
        os.system('cls')
        print(" 1 - Generar cartas")
        print(" 2 - Salir ")
        opcion = int(input("Ingrese una opción: "))
        valida_opcion(opcion)
        if opcion == 1:
            os.system('cls')
            combinar_correspondencia()
            pass
            os.system('pause')
        elif opcion == 2:
            os.system("cls")
            print('Usted esta saliendo del programa. ¡Hasta pronto!')
            os.system("pause")
        else: 
            print("La opción ingresada no es válida")

def main():
    while True:
        try: 
            os.system("cls")
            menu()
            break
        except ValueError as e:
            print(e)
        except AssertionError as e:
            print(e)
        finally:
            pass

if __name__=='__main__':
    main()