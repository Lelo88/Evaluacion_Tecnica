import openpyxl

book1=openpyxl.load_workbook('PlanilladeHoras-CoordinadorA.xlsx', data_only=True)
book2=openpyxl.load_workbook('PlanilladeHoras-CoordinadorB.xlsx', data_only=True)
book3=openpyxl.load_workbook('PlanilladeHoras-CoordinadorC.xlsx', data_only=True)
book4=openpyxl.load_workbook('PlanilladeHoras-CoordinadorD.xlsx', data_only=True)

bookFinal = openpyxl.Workbook()

hoja1 = book1.active
hoja2 = book2.active
hoja3 = book3.active
hoja4 = book4.active

fila4 = []
for fila in hoja4.iter_rows(min_row=2):
    dato = ([celda.value for celda in fila])
    fila4.append(dato)

for fila in fila4:
    hoja3.append(fila)

fila3 = []
for fila in hoja3.iter_rows(min_row=2):
    dato = ([celda.value for celda in fila])
    fila3.append(dato)
    
for fila in fila3:
    hoja2.append(fila)    

fila2 = []
for fila in hoja2.iter_rows(min_row=2):
    dato = ([celda.value for celda in fila])
    fila2.append(dato)

for fila in fila2:
    hoja1.append(fila)

book1.save('EjercicioFinal.xlsx')